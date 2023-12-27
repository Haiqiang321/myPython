# _*_ coding:utf-8 _*_

# @Author :Haiqiang
# @Time :2023/10/25 11:02
# @FileName :demo.py

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelHandle:
    """
    操作excel
    """
    def __init__(self, file):
        self.file = file

    def open_sheet(self, sheet_name) -> Worksheet:
        """
        打开表单
        """
        wb = load_workbook(self.file)
        sheet = wb[sheet_name]
        return sheet

    def read_header(self, sheet_name):
        """
        获取表头
        :param sheet_name:
        :return:
        """
        sheet = self.open_sheet(sheet_name)
        headers = []
        for i in sheet[1]:
            headers.append(i.value)
        return headers

    def read_rows(self, sheet_name):
        """
        读取除表头外所有数据，（除第一行外的所有数据）
        :param sheet_name:
        :return:
        """
        sheet = self.open_sheet(sheet_name)
        rows = list(sheet.rows)[1:]

        data = []
        for row in rows:
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            data.append(row_data)

        return data

    def read_key_value(self, sheet_name):
        """
        获取所有数据，且将表头中的内容与数据结合展示（以字典的形式）
        :param sheet_name:
        :return:
        """
        sheet = self.open_sheet('Sheet1')
        rows = list(sheet.rows)

        # 获取标题
        data = []
        for row in rows[1:]:
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            data_dict = dict(zip(self.read_header(sheet_name), row_data))
            data.append(data_dict)
        return data

    @staticmethod
    def write_change(file, sheet_name, row, colum, data):
        """
        写入Excel 数据
        :param file:
        :param sheet_name:
        :param row:
        :param colum:
        :param data:
        :return:
        """
        wb = load_workbook(file)
        sheet = wb[sheet_name]

        sheet.cell(row, colum).value = data
        wb.save(file)
        wb.close()


if __name__ == '__main__':
    file = "D:\\workshop\\code\\python\\mypython\\test.xlsx"
    excel_file = ExcelHandle(file)

    sheet = excel_file.open_sheet('Sheet1')
    change_value = excel_file.write_change('test.xlsx', 'Sheet1', 2, 3, '0')


